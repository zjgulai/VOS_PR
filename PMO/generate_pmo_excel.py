"""生成 PMO 业务协作 Excel（PR 团队 / 社媒团队 分开，AI 预填参考答案 + 选择原因 + 优先级）。

用法：python3 PMO/generate_pmo_excel.py
输出：PMO/业务协作_社媒团队/社媒团队协作表.xlsx
      PMO/业务协作_PR团队/PR团队协作表.xlsx
约定：白色列 = AI 预填（参考答案 + 原因 + 优先级），黄色「业务确认」列 = 业务填 确认/修改/删除。
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path.home() / "Library/Python/3.9/lib/python/site-packages"))
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from tools.social.dictionary import load_dictionary

PROJ = Path(__file__).resolve().parents[1]
PMO = PROJ / "PMO"

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
FILL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP = Alignment(vertical="top", wrap_text=True)


def _header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def _fill_col(ws, col: int, nrows: int) -> None:
    for r in range(2, nrows + 2):
        ws.cell(row=r, column=col).fill = FILL_FILL


def _widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _append_sheet(ws, headers: list[str], rows: list[list], confirm_col: int,
                  widths: list[int], blank_rows: int = 3) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row + [""] * (len(headers) - len(row)))
    for _ in range(blank_rows):
        ws.append([""] * len(headers))
    _header(ws, len(headers))
    nrows = ws.max_row - 1
    _fill_col(ws, confirm_col, nrows)
    for c in range(1, len(headers) + 1):
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=c).alignment = WRAP
    _widths(ws, widths)
    ws.freeze_panes = "A2"


# ───────────────────────── 社媒团队数据 ─────────────────────────

SOCIAL_GROUPS = [
    ["Exclusively Pumping Mamas - Education & Support Group",
     "https://www.facebook.com/groups/1574856819503023/",
     "排奶/泵奶支持", "教育+支持型，S1 用户讨论核心群，排奶人群高度相关", "P0"],
    ["Exclusive Pumping: Breastfeeding Without Nursing",
     "https://www.facebook.com/groups/EP.BWN/",
     "排奶/泵奶支持", "经典排奶大群，跨地区，话题密度高", "P0"],
    ["Wearable Pump Paperweight Prevention",
     "URL 待补充", "可穿戴泵 troubleshooting",
     "IBCLC Jessica Anderson 运营，可穿戴泵专业讨论，与 Momcozy 品类直接相关", "P0"],
    ["Breastfeeding Support Group for Black Moms",
     "URL 待补充", "母乳支持",
     "约 9.1 万成员大群，多元人群覆盖，可观察不同人群需求", "P1"],
    ["Pumping Mamas: breastfeeding support for working moms",
     "URL 待补充", "职场泵奶",
     "职场妈妈泵奶场景，与 Momcozy 核心目标人群（职场泵奶）重合", "P1"],
]

SOCIAL_KOLS = [
    ["Allison Tolman / New Little Life", "YouTube", "泵奶器测评 IBCLC",
     "头部泵奶器专家，可做专家背书/Pitch，深度测评", "P0"],
    ["Jessica Anderson / Genuine Lactation", "Instagram + FB", "可穿戴泵 IBCLC",
     "可穿戴泵专业测评，运营可穿戴泵社群", "P0"],
    ["Allison Banfield / Pumping Milk", "YouTube + Blog", "Momcozy 测评",
     "Momcozy 全系列测评（M5/M6/M9/V1 Pro），可直接合作", "P0"],
    ["Carisa Myers / The Milk Effect", "Instagram", "泵奶器 RN IBCLC",
     "测评 30+ 泵，100 分评分体系，无品牌赞助，客观", "P1"],
    ["Katie Clark / The Breastfeeding Mama", "YouTube", "泵奶器 IBCLC",
     "IBCLC 推荐类内容，6.7K+ 播放", "P1"],
    ["Tamari Jacob / @onewiththepump", "Instagram / Threads", "排奶/泵奶",
     "16.9K 粉丝，排奶教育", "P1"],
    ["maxinesanz", "Instagram", "泵奶器测评",
     "新手妈妈泵奶测评，#momcozy #spectra", "P2"],
]


def _social_competitor_rows() -> list[list]:
    d = load_dictionary()
    fb = d.get("facebook_pages", {})
    rows = []
    for line in d["competitors"]["pump"] + d["competitors"]["feeding_appliance"]:
        pl = "吸奶器" if line in d["competitors"]["pump"] else "喂养电器"
        rows.append([
            line["name"], pl, line["tier"], line.get("tiktok") or "",
            line.get("instagram") or "", fb.get(line["brand_key"], ""),
            f"T{line['tier']} {line['priority']} 直接竞品，建议监测", line["priority"],
        ])
    return rows


def gen_social_excel() -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "1.FacebookGroups群组"
    _append_sheet(ws,
        ["群组名称", "群组URL", "主题", "AI推荐原因", "优先级", "业务确认"],
        SOCIAL_GROUPS, 6, [40, 40, 20, 40, 8, 12])

    ws = wb.create_sheet("2.KOL关注池")
    _append_sheet(ws,
        ["Creator账号", "平台", "垂类", "AI推荐原因", "优先级", "业务确认"],
        SOCIAL_KOLS, 6, [32, 16, 18, 42, 8, 12])

    ws = wb.create_sheet("3.竞品社媒账号")
    _append_sheet(ws,
        ["品牌", "品线", "层级", "TikTok", "Instagram", "Facebook", "AI推荐原因", "优先级", "业务确认"],
        _social_competitor_rows(), 9, [14, 10, 6, 18, 20, 30, 22, 8, 12])

    ws = wb.create_sheet("4.周报接收人")
    _append_sheet(ws,
        ["角色", "姓名/账号", "周报语言", "接收方式", "AI参考建议", "业务确认"],
        [
            ["社媒 Analyst", "", "中/英待定", "飞书/邮件待定",
             "建议设专职 Analyst 每周审核 30-60 分钟后发布", ""],
            ["社媒内容负责人", "", "中/英待定", "飞书/邮件待定",
             "建议同步周报给内容负责人，便于选题落地", ""],
        ], 6, [18, 16, 14, 16, 40, 12])

    out = PMO / "业务协作_社媒团队" / "社媒团队协作表.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"✓ {out.relative_to(PROJ)} (4 sheets)")


# ───────────────────────── PR 团队数据 ─────────────────────────

PR_MEDIA = [
    ["What to Expect", "缺席", "两份吸奶器榜单均无 Momcozy",
     "最高优先缺口，需送测刷新", "P0"],
    ["Wirecutter (NYT)", "弱负面", "S9 Pro 被批 'clunky'/'90s pager'",
     "送新品重测是唯一修复路径", "P0"],
    ["Forbes Vetted", "强", "M5 = Best Value",
     "当前最强单一媒体关系节点，Alicia Betz", "P0"],
    ["Consumer Reports", "强", "4 款入测",
     "non-WiFi 隐私卖点契合 CR 隐私议程", "P0"],
    ["Babylist", "强", "M5 = Best Affordable",
     "对比页为赞助内容，竞品可质疑，需维护", "P1"],
    ["The Bump", "强", "V1 Pro = 最佳整体",
     "频繁重排，至少提前 8 周备机", "P1"],
    ["Reviewed (USA TODAY)", "缺席", "常青榜全部竞品占位",
     "高优先缺口，需攻入", "P1"],
    ["Good Housekeeping", "缺席", "GH Institute 独立实验室",
     "英国版曾给 M6 93/100，美国版需攻入测试池", "P1"],
    ["Modern Retail", "负面", "已将 Momcozy 列为 dupe/跟随者",
     "需主动纠正叙事", "P1"],
    ["BabyGearLab", "中", "深度参数化测评",
     "吸力/噪声量化，适合技术叙事", "P2"],
    ["Parents", "中", "Best for Baby 2026 年度奖",
     "获奖明细需内部确认", "P2"],
    ["TIME", "弱", "Best Inventions 年度榜",
     "竞品 Frida 2025 已入选，需跟进", "P2"],
]

PR_EXPERTS = [
    ["Allison Tolman", "泵奶器 IBCLC", "可对外背书",
     "New Little Life 创始人，泵奶器专家", "P0"],
    ["Jessica Anderson", "可穿戴泵 IBCLC", "可对外背书",
     "Genuine Lactation，可穿戴泵专业", "P0"],
]

PR_MEDIA_CONTACTS = [
    ["Sarah Connor", "bestproducts.com", "1M/月", "免费（寄样+挂亚马逊链接）",
     "建议合作：Momcozy 用户且孕期 10 周，测评/礼物指南", "P0"],
    ["Hana Amaes", "Medical News Today", "17M/月", "医学 blog 500欧/1500字",
     "建议推进：大流量，有孕哺护肤内容经验", "P0"],
    ["Kate / Baby Chick", "baby-chick.com", "900万曝光/月", "$7,500-$15,000",
     "建议合作：孕期/新手妈妈受众，组合套餐", "P0"],
    ["Susey Harmer", "Today's Parent", "600万读者/月", "$18,500 / $11,000",
     "待确认：权威强但价高，需曝光/点击数据", "P1"],
    ["Tiffany Tse", "SheKnows", "8000万+/月", "待询价",
     "待判断：需明确报道形式/是否挂链接/排期", "P1"],
    ["Kari Bliss", "Sophisticated Whimsy", "未提供", "$1,500",
     "待判断：6个月新手妈妈，需补媒体资料包", "P1"],
    ["Sydney Wingfield", "The Skin (Substack)", "400-1000 views", "$300",
     "小预算尝试：skincare 场景，曝光有限", "P2"],
    ["Angela Onwuzoo", "healthwise.punchng", "108K/月", "—",
     "不合作：尼日利亚受众，不相关", "P2"],
    ["Hannah Payne", "threadsbyh.com", "无数据", "—",
     "不合作：旅游方向，不匹配", "P2"],
]

PR_RULES = [
    ["媒体 ROI 过滤规则", "月流量 <100万 且 报价 >$2,000 → 不推荐",
     "工作日志筛掉 Angela Onwuzoo(108K)、Sydney Wingfield(400views) 的经验固化", "P0"],
    ["保障曝光先问清", "凡'保障报道/曝光'但不明确是否挂链接/赞助标注/追踪链接 → 先问清再判断",
     "工作日志 Tiffany Tse(SheKnows) 待判断的经验", "P0"],
    ["媒体判断 5 维度", "必查：网站流量、受众地域、内容方向匹配、报价合理性、可追踪性",
     "工作日志实操标准", "P0"],
    ["海外 Earned Media 逻辑", "海外媒体不可'买版面保出稿'，编辑决定是否报道，新闻价值>品牌需求",
     "一文吃透PR全貌（跨境 PR 底层逻辑）", "P0"],
    ["Pitch 个性化 + 冷却期", "一对一 Pitch，群发模板会被编辑拉黑；同一编辑 pitch 频率过高需冷却",
     "海外PR vs 国内公关（避坑）", "P1"],
    ["PR 预算框架", "通稿月度 3 篇、单篇 ~$2,500、月度 6-8k；外链 $1,500/月",
     "工作日志预算记录，需确认是否延续", "P1"],
]


def gen_pr_excel() -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "1.核心媒体与编辑"
    _append_sheet(ws,
        ["媒体", "当前关系", "关键机会/风险", "AI推荐原因", "优先级", "业务确认"],
        PR_MEDIA, 6, [22, 10, 34, 34, 8, 12])

    ws = wb.create_sheet("2.周报接收人")
    _append_sheet(ws,
        ["角色", "姓名/账号", "周报语言", "接收方式", "AI参考建议", "业务确认"],
        [
            ["Global PR Lead", "", "中/英待定", "飞书/邮件待定",
             "建议每周 30 分钟看周报掌握全局", ""],
            ["PR Analyst", "", "中/英待定", "飞书/邮件待定",
             "建议设专职 Analyst 审核周报后发布", ""],
            ["Regional PR", "", "中/英待定", "飞书/邮件待定",
             "建议每日看 P0/P1 预警及时响应本地风险", ""],
        ], 6, [18, 16, 14, 16, 40, 12])

    ws = wb.create_sheet("3.专家与样机库")
    _append_sheet(ws,
        ["专家/资源", "领域", "可对外", "AI参考建议", "优先级", "业务确认"],
        PR_EXPERTS, 6, [22, 16, 12, 40, 8, 12])

    ws = wb.create_sheet("4.媒体联系人库")
    _append_sheet(ws,
        ["联系人", "媒体", "月流量", "报价", "AI合作结论", "优先级", "业务确认"],
        PR_MEDIA_CONTACTS, 7, [18, 24, 14, 22, 42, 8, 12])

    ws = wb.create_sheet("5.规则确认")
    _append_sheet(ws,
        ["规则名称", "规则内容", "AI推荐依据", "优先级", "业务确认", "业务修订"],
        PR_RULES, 5, [20, 50, 40, 8, 12, 20])

    out = PMO / "业务协作_PR团队" / "PR团队协作表.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"✓ {out.relative_to(PROJ)} (5 sheets)")


if __name__ == "__main__":
    gen_social_excel()
    gen_pr_excel()
    print("完成")
