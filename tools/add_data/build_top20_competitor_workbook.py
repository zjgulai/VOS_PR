from __future__ import annotations

import argparse
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

PROJ = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJ))

from tools.cleaning._common import BRAND_NORMALIZE_MAP, get_code

TABLES_DIR = PROJ / "data" / "delivery" / "tables"
DEFAULT_OUT_DIR = PROJ / "data" / "add_data"
OWN_BRANDS = {"momcozy"}

KNOWN_BRANDS = {
    "Ameda",
    "Ardo",
    "Baby Brezza",
    "Baby Buddha",
    "BabyBjorn",
    "Babycare",
    "BabyOno",
    "Babymoov",
    "Bebefun",
    "Béaba",
    "Bébé Confort",
    "BellaBaby",
    "Britax",
    "Bugaboo",
    "Canpol babies",
    "Chicco",
    "Cosatto",
    "Cybex",
    "Dr. Brown's",
    "Elvie",
    "Evenflo",
    "Eufy",
    "Goodbaby",
    "Graco",
    "Haakaa",
    "Hartan",
    "Hauck",
    "iCandy",
    "Inglesina",
    "Jané",
    "Joie",
    "Joolz",
    "Kinderkraft",
    "Lacteck",
    "Lansinoh",
    "LOVI",
    "MAM",
    "Medela",
    "Miniland",
    "Motif Medical",
    "Momcozy",
    "Mumma Bump",
    "New Beginnings",
    "NUK",
    "Peg Perego",
    "Philips Avent",
    "Pigeon",
    "reer",
    "Redsbaby",
    "Silver Cross",
    "Spectra",
    "Stokke",
    "Suavinex",
    "Tommee Tippee",
    "UPPAbaby",
    "Willow",
    "Xiao Bai Xiong",
    "小白熊",
    "好孩子",
}

BRAND_SITE_MAP = {
    "Ameda": "https://www.ameda.com/",
    "Ardo": "https://www.ardomedical.com/",
    "Baby Brezza": "https://babybrezza.com/",
    "Baby Buddha": "https://babybuddhaproducts.com/",
    "BabyBjorn": "https://www.babybjorn.com/",
    "Babycare": "https://www.babycare.com/",
    "BabyOno": "https://babyono.com/en/about-us/10",
    "Babymoov": "https://www.babymoov.fr/",
    "Béaba": "https://www.beaba.com/",
    "Bébé Confort": "https://www.bebeconfort.com/",
    "Britax": "https://us.britax.com/",
    "Bugaboo": "https://www.bugaboo.com/",
    "Canpol babies": "https://canpolbabies.com/",
    "Chicco": "https://www.chicco.com/",
    "Cosatto": "https://www.cosatto.com/",
    "Cybex": "https://www.cybex-online.com/",
    "Dr. Brown's": "https://drbrownsbaby.com/",
    "Elvie": "https://www.elvie.com/",
    "Evenflo": "https://www.evenflofeeding.com/",
    "Eufy": "https://www.eufy.com/",
    "Goodbaby": "https://www.gbinternational.com.hk/",
    "Graco": "https://www.gracobaby.com/",
    "Haakaa": "https://haakaa.co.nz/",
    "Hartan": "https://www.hartan.de/",
    "Hauck": "https://hauck.de/",
    "iCandy": "https://www.icandyworld.com/",
    "Inglesina": "https://www.inglesina.com/",
    "Jané": "https://www.jane.es/es/",
    "Joie": "https://joiebaby.com/",
    "Joolz": "https://www.joolz.com/",
    "Kinderkraft": "https://kinderkraft.pl/",
    "Lansinoh": "https://lansinoh.com/",
    "LOVI": "https://lovi.pl/",
    "MAM": "https://www.mambaby.com/",
    "Medela": "https://www.medela.com/",
    "Miniland": "https://web.minilandgroup.com/",
    "Motif Medical": "https://motifmedical.com/",
    "New Beginnings": "https://newbeginnings.com.au/",
    "NUK": "https://www.nuk.de/",
    "Peg Perego": "https://www.pegperego.com/",
    "Philips Avent": "https://www.philips.com/c-e/mo/philips-avent-parenting.html",
    "Pigeon": "https://www.pigeon.com/",
    "reer": "https://www.reer.de/",
    "Redsbaby": "https://www.redsbaby.com/en-au/why-redsbaby",
    "Silver Cross": "https://www.silvercrossbaby.com/",
    "Spectra": "https://www.spectrababyusa.com/",
    "Stokke": "https://www.stokke.com/",
    "Suavinex": "https://www.suavinex.com/en/",
    "Tommee Tippee": "https://www.tommeetippee.com/",
    "UPPAbaby": "https://uppababy.com/",
    "Willow": "https://onewillow.com/",
    "Zomee": "https://zomee.com/",
    "小白熊": "https://www.snow-bear.cn/brand/detail/id/378.html",
    "好孩子": "https://www.gbinternational.com.hk/",
}

GENERIC_PATTERNS = [
    r"待业务",
    r"请按",
    r"以.*为准",
    r"品类参考",
    r"无有效型号",
    r"无单一商业品牌",
    r"多品牌",
    r"文章列表",
    r"专题",
    r"列表",
    r"栏目",
    r"检索",
    r"站内",
    r"结果",
    r"结果列表",
    r"聚合",
    r"类目",
    r"链接",
    r"review",
    r"guide",
    r"comparison",
    r"worth it",
    r"manual$",
    r"electric pump$",
    r"wearable$",
    r"bottle warmer$",
    r"sterilizer$",
    r"travel system$",
    r"stroller$",
    r"breast pump$",
    r"extractor de leche$",
    r"sacaleches$",
    r"tire-lait$",
    r"milchpumpe$",
    r"شفاط حليب$",
    r"carriola$",
    r"cochecito$",
    r"poussette$",
    r"kinderwagen$",
    r"电动泵$",
    r"手动泵$",
    r"手动$",
    r"电动$",
]

MODEL_HINTS = {
    "Advanced Double Electric",
    "Aura Glow",
    "Calypso",
    "Freestyle",
    "Harmony",
    "Magic InBra",
    "Pump in Style",
    "S1",
    "S1+",
    "S2",
    "S2+",
    "S2 Plus",
    "Shelly",
    "Swing",
    "Swing Maxi",
    "Willow Go",
}

WEB_SUPPLEMENTS = [
    {
        "country": "英国",
        "product_lines": ["家居出行"],
        "local_brands": ["Silver Cross", "iCandy", "Cosatto"],
        "query": "英国 本土 婴儿 推车 品牌 Silver Cross iCandy Cosatto Tommee Tippee 官方",
        "urls": [
            "https://www.silvercrossbaby.com/",
            "https://cn.silvercrossbaby.com/",
            "https://ukbabycentre.com/",
        ],
        "note": "检索结果显示 Silver Cross 为英国原创品牌，iCandy/Cosatto 作为英国本土推车品牌补充写入。",
    },
    {
        "country": "英国",
        "product_lines": ["喂养电器"],
        "local_brands": ["Tommee Tippee"],
        "query": "英国 本土 婴儿 推车 品牌 Silver Cross iCandy Cosatto Tommee Tippee 官方",
        "urls": [
            "https://www.tommeetippee.com/",
        ],
        "note": "Tommee Tippee 作为英国本土喂养品牌补充写入。",
    },
    {
        "country": "法国",
        "product_lines": ["喂养电器", "智能母婴电器"],
        "local_brands": ["Babymoov", "Béaba"],
        "query": "法国 本土 母婴 品牌 Babymoov Beaba Bebe Confort 官方",
        "urls": [
            "https://www.babymoov.fr/",
            "http://beaba.com.cn/?website%2F=",
        ],
        "note": "Babymoov、Béaba 为法国母婴品牌，适合补喂养/辅食/智能设备相关本土品牌位。",
    },
    {
        "country": "法国",
        "product_lines": ["家居出行"],
        "local_brands": ["Bébé Confort"],
        "query": "法国 本土 母婴 品牌 Babymoov Beaba Bebe Confort 官方",
        "urls": [
            "https://www.lesbebesdubonheur.com/brands/B%C3%A9b%C3%A9-Confort",
        ],
        "note": "Bébé Confort 用于补法国本土出行类品牌。",
    },
    {
        "country": "西班牙",
        "product_lines": ["喂养电器", "智能母婴电器"],
        "local_brands": ["Suavinex", "Miniland"],
        "query": "西班牙 本土 母婴 品牌 Suavinex Jane Miniland 官方",
        "urls": [
            "https://www.suavinex.com/en/about-us",
            "https://web.minilandgroup.com/",
        ],
        "note": "Suavinex、Miniland 为西班牙本土母婴品牌，适合补喂养/监测类。",
    },
    {
        "country": "西班牙",
        "product_lines": ["家居出行"],
        "local_brands": ["Jané"],
        "query": "西班牙 Jané 官方 品牌 about us",
        "urls": [
            "http://groupjane.com/es",
            "http://www.jane.es/es/",
        ],
        "note": "Jané 为西班牙本土婴童出行品牌。",
    },
    {
        "country": "意大利",
        "product_lines": ["家居出行"],
        "local_brands": ["Peg Perego", "Inglesina", "Chicco"],
        "query": "意大利 本土 母婴 品牌 Chicco Peg Perego Inglesina 官方",
        "urls": [
            "http://www.chicco.com/it.html",
            "https://www.pegperego.com/",
            "https://www.inglesina.com/",
        ],
        "note": "Peg Perego、Inglesina、Chicco 为意大利本土出行品牌补充。",
    },
    {
        "country": "意大利",
        "product_lines": ["喂养电器", "吸奶器"],
        "local_brands": ["Chicco"],
        "query": "意大利 本土 母婴 品牌 Chicco Peg Perego Inglesina 官方",
        "urls": [
            "http://www.chicco.com/it.html",
        ],
        "note": "Chicco 适合补意大利本土喂养/吸奶器品牌位。",
    },
    {
        "country": "波兰",
        "product_lines": ["家居出行"],
        "local_brands": ["Kinderkraft"],
        "query": "波兰 本土 母婴 品牌 Kinderkraft Canpol BabyOno LOVI 官方",
        "urls": [
            "https://kinderkraft.pl/",
        ],
        "note": "Kinderkraft 为波兰本土推车/汽座品牌。",
    },
    {
        "country": "波兰",
        "product_lines": ["喂养电器", "吸奶器", "智能母婴电器"],
        "local_brands": ["Canpol babies", "LOVI", "BabyOno"],
        "query": "BabyOno 官方 波兰 品牌",
        "urls": [
            "http://www.babyono.pl/",
            "https://babyono.com/en/about-us/10",
            "https://www.krak-wit.pl/pl/c/CIAZA-I-MACIERZYNSTWO/341/2/desc/3/promotion/0/f_producer_44/1/f_producer_136/1",
        ],
        "note": "BabyOno、Canpol babies、LOVI 作为波兰本土喂养/吸奶器品牌补充。",
    },
    {
        "country": "德国",
        "product_lines": ["家居出行"],
        "local_brands": ["Hauck", "Hartan"],
        "query": "德国 本土 母婴 品牌 Hauck Hartan NUK reer 官方",
        "urls": [
            "https://hauck.de/",
        ],
        "note": "Hauck 为德国本土母婴出行品牌；Hartan 作为德国推车品牌补充写入。",
    },
    {
        "country": "德国",
        "product_lines": ["喂养电器", "吸奶器"],
        "local_brands": ["NUK", "reer"],
        "query": "德国 本土 母婴 品牌 Hauck Hartan NUK reer 官方",
        "urls": [
            "https://www.nuk.de/",
            "https://www.reer.de/",
        ],
        "note": "NUK、reer 作为德国本土喂养/护理品牌补充。",
    },
    {
        "country": "荷兰",
        "product_lines": ["家居出行"],
        "local_brands": ["Bugaboo", "Joolz"],
        "query": "荷兰 本土 婴儿 推车 品牌 Bugaboo Joolz 官方",
        "urls": [
            "https://www.bugaboo.com/nl-nl",
            "https://www.joolz.com/nl/nl/home",
        ],
        "note": "Bugaboo、Joolz 为荷兰本土推车品牌。",
    },
    {
        "country": "奥地利",
        "product_lines": ["喂养电器", "吸奶器"],
        "local_brands": ["MAM"],
        "query": "奥地利 本土 母婴 品牌 MAM 官方",
        "urls": [
            "https://www.mambaby.com/at/warum-mam/ueber-mam/",
        ],
        "note": "MAM 为奥地利本土母婴品牌，适合补喂养/哺乳相关位。",
    },
    {
        "country": "澳大利亚",
        "product_lines": ["家居出行"],
        "local_brands": ["Redsbaby"],
        "query": "澳大利亚 本土 婴儿 推车 品牌 Redsbaby 官方 母婴",
        "urls": [
            "https://www.redsbaby.com/en-au/why-redsbaby",
            "https://www.redsbaby.com/en-au/shop/prams-strollers",
        ],
        "note": "Redsbaby 为澳大利亚本土推车品牌。",
    },
    {
        "country": "中国",
        "product_lines": ["喂养电器", "智能母婴电器", "吸奶器"],
        "local_brands": ["小白熊", "Babycare"],
        "query": "中国 本土 母婴 品牌 Babycare Goodbaby 小白熊 官方",
        "urls": [
            "https://www.snow-bear.cn/brand/detail/id/378.html",
            "https://www.babycare.com/",
        ],
        "note": "小白熊、Babycare 适合补中国本土喂养/母婴电器品牌。",
    },
    {
        "country": "中国",
        "product_lines": ["家居出行"],
        "local_brands": ["Goodbaby", "好孩子"],
        "query": "中国 本土 母婴 品牌 Babycare Goodbaby 小白熊 官方",
        "urls": [
            "https://www.gbinternational.com.hk/",
        ],
        "note": "Goodbaby/好孩子 适合补中国本土出行品牌。",
    },
]

NORMALIZE_EXTRA = {
    "avent": "Philips Avent",
    "baby ono": "BabyOno",
    "béaba": "Béaba",
    "bébé confort": "Bébé Confort",
    "canpol": "Canpol babies",
    "goodbaby": "Goodbaby",
    "haakaa": "Haakaa",
    "icandy": "iCandy",
    "joolz": "Joolz",
    "kinderkraft": "Kinderkraft",
    "lovi": "LOVI",
    "mam": "MAM",
    "miniland": "Miniland",
    "nuk": "NUK",
    "redsbaby": "Redsbaby",
    "silver cross": "Silver Cross",
    "snowbear": "小白熊",
    "suavinex": "Suavinex",
    "tommee tippee": "Tommee Tippee",
    "xiaobaixiong": "小白熊",
    "小白熊": "小白熊",
    "好孩子": "好孩子",
}


def clean_text(v: object) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return str(v).replace("\xa0", " ").strip()


def normalize_brand_name(v: str) -> str:
    s = clean_text(v)
    if not s:
        return ""
    s = re.sub(r"\s+", " ", s).strip(" ,;|/·-")
    key = s.lower()
    if key in BRAND_NORMALIZE_MAP:
        return BRAND_NORMALIZE_MAP[key]
    if key in NORMALIZE_EXTRA:
        return NORMALIZE_EXTRA[key]
    for brand in sorted(KNOWN_BRANDS, key=len, reverse=True):
        if key == brand.lower():
            return brand
    return s


def base_brand_from_token(token: str) -> str:
    cleaned = clean_text(token)
    cleaned = re.sub(r"[（(].*?[)）]", "", cleaned).strip()
    cleaned = cleaned.strip(" ,;|/·-")
    cleaned = normalize_brand_name(cleaned)
    for brand in sorted(KNOWN_BRANDS, key=len, reverse=True):
        if cleaned.lower() == brand.lower():
            return brand
        if cleaned.lower().startswith(brand.lower() + " "):
            return brand
    return cleaned


def is_own_brand(token: str) -> bool:
    return base_brand_from_token(token).lower() in OWN_BRANDS


def looks_generic(token: str) -> bool:
    s = clean_text(token)
    if not s:
        return True
    lowered = s.lower()
    for pat in GENERIC_PATTERNS:
        if re.search(pat, lowered, flags=re.I):
            return True
    if len(re.sub(r"[\W_]+", "", lowered)) <= 1:
        return True
    return False


def looks_like_model(token: str) -> bool:
    s = clean_text(token)
    if not s:
        return False
    compact = re.sub(r"[（(].*?[)）]", "", s).strip()
    if any(h.lower() in compact.lower() for h in MODEL_HINTS):
        return True
    if re.search(r"\b[A-Z]?\d+[A-Z+\-]*\b", compact, flags=re.I):
        return True
    brand = base_brand_from_token(compact)
    if brand and compact.lower().startswith(brand.lower() + " "):
        tail = compact[len(brand):].strip(" -")
        if tail and not looks_generic(tail):
            return True
    return False


def tokenize_block(text: str) -> list[str]:
    if not text:
        return []
    text = re.sub(r"【.*?】", "", text)
    text = text.replace("；", ";").replace("，", ",").replace("、", ",").replace("｜", "|")
    text = text.replace("\n", ",")
    parts = [p.strip() for p in re.split(r"[,;|]", text) if p.strip()]
    return parts


def split_brand_sections(raw: str) -> tuple[str, str, str]:
    s = clean_text(raw)
    if not s:
        return "", "", ""
    s = re.sub(r"【.*?】", "", s).strip()
    m = re.search(r"国际[^:：]*[:：](.*?)(?:[;；]|\|)\s*本土(?:/其他)?[^:：]*[:：](.*)$", s, flags=re.I)
    if m:
        return m.group(1).strip(), m.group(2).strip(), ""
    m = re.search(r"国际[^:：]*[:：](.*)$", s, flags=re.I)
    intl = m.group(1).strip() if m else ""
    m2 = re.search(r"本土(?:/其他)?[^:：]*[:：](.*)$", s, flags=re.I)
    local = m2.group(1).strip() if m2 else ""
    if intl or local:
        return intl, local, ""
    return "", "", s


def dedupe_keep_order(items: list[str]) -> list[str]:
    seen = set()
    out = []
    for x in items:
        k = clean_text(x)
        if not k:
            continue
        if k.lower() in seen:
            continue
        seen.add(k.lower())
        out.append(k)
    return out


def classify_tokens(country: str, product_line: str, raw_brands: str, raw_models: str) -> tuple[dict, list[dict]]:
    intl_block, local_block, residual = split_brand_sections(raw_brands)
    audit_rows: list[dict] = []
    result = {
        "intl_brands": [],
        "local_brands": [],
        "intl_models": [],
        "local_models": [],
        "misc_notes": [],
        "classification_note": [],
        "removed_own_brand": [],
    }

    def handle(token: str, source_field: str, preferred_bucket: str | None) -> None:
        token = clean_text(token)
        if not token:
            return
        token_wo_note = re.sub(r"[（(].*?[)）]", "", token).strip()
        normalized_brand = base_brand_from_token(token_wo_note)
        row = {
            "country": country,
            "product_line": product_line,
            "source_field": source_field,
            "raw_token": token,
            "normalized_brand": normalized_brand,
            "classification": "",
            "final_value": "",
            "note": "",
        }
        if is_own_brand(token_wo_note):
            row["classification"] = "excluded_own_brand"
            row["final_value"] = normalized_brand or token_wo_note
            result["removed_own_brand"].append(normalized_brand or token_wo_note)
            audit_rows.append(row)
            return
        if looks_generic(token_wo_note):
            row["classification"] = "generic_or_note"
            row["final_value"] = token_wo_note
            result["misc_notes"].append(token_wo_note)
            audit_rows.append(row)
            return
        if source_field == "raw_models" or looks_like_model(token_wo_note):
            model_value = token_wo_note
            if normalized_brand and normalized_brand in KNOWN_BRANDS and model_value.lower() == normalized_brand.lower():
                if preferred_bucket == "local":
                    result["local_brands"].append(normalized_brand)
                elif preferred_bucket == "intl":
                    result["intl_brands"].append(normalized_brand)
                else:
                    result["intl_brands"].append(normalized_brand)
                row["classification"] = "brand"
                row["final_value"] = normalized_brand
                audit_rows.append(row)
                return
            model_bucket = preferred_bucket
            if model_bucket is None and normalized_brand:
                if normalized_brand in result["local_brands"]:
                    model_bucket = "local"
                elif normalized_brand in result["intl_brands"]:
                    model_bucket = "intl"
            if model_bucket == "local":
                result["local_models"].append(model_value)
            else:
                result["intl_models"].append(model_value)
            row["classification"] = f"model_{model_bucket or 'intl'}"
            row["final_value"] = model_value
            audit_rows.append(row)
            return
        bucket = preferred_bucket or "intl"
        if bucket == "local":
            result["local_brands"].append(normalized_brand)
        else:
            result["intl_brands"].append(normalized_brand)
        row["classification"] = f"brand_{bucket}"
        row["final_value"] = normalized_brand
        audit_rows.append(row)

    for token in tokenize_block(intl_block):
        handle(token, "raw_brands", "intl")
    for token in tokenize_block(local_block):
        handle(token, "raw_brands", "local")
    for token in tokenize_block(residual):
        handle(token, "raw_brands", None)
    for token in tokenize_block(raw_models):
        handle(token, "raw_models", None)

    result["intl_brands"] = dedupe_keep_order([normalize_brand_name(x) for x in result["intl_brands"] if x])
    result["local_brands"] = dedupe_keep_order([normalize_brand_name(x) for x in result["local_brands"] if x])
    result["intl_models"] = dedupe_keep_order(result["intl_models"])
    result["local_models"] = dedupe_keep_order(result["local_models"])
    result["misc_notes"] = dedupe_keep_order(result["misc_notes"])

    models_lower = {m.lower() for m in result["intl_models"] + result["local_models"]}
    result["intl_brands"] = [b for b in result["intl_brands"] if b.lower() not in models_lower]
    result["local_brands"] = [b for b in result["local_brands"] if b.lower() not in models_lower]
    if intl_block and local_block:
        result["classification_note"].append("原始字段已显式区分国际/本土")
    elif raw_brands:
        result["classification_note"].append("原始字段未完全区分国际/本土，未明确项默认放入国际/待审池")
    if result["removed_own_brand"]:
        result["classification_note"].append("已从竞品列排除自有品牌 Momcozy")
    return result, audit_rows


def build_voc_top_brands(voc_df: pd.DataFrame, top20_countries: set[str]) -> dict[tuple[str, str], str]:
    if voc_df.empty:
        return {}
    df = voc_df.copy()
    rename_map = {
        "国家": "country",
        "产品品线": "product_line",
        "竞品关联品牌": "competitor_brand",
        "批次编码": "batch_code",
    }
    df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})
    df["country"] = df["country"].map(clean_text)
    df["product_line"] = df["product_line"].map(clean_text)
    df["competitor_brand"] = df["competitor_brand"].map(clean_text)
    df["batch_code"] = df["batch_code"].map(clean_text)
    df = df[df["country"].isin(top20_countries)]
    df = df[~df["batch_code"].str.startswith("MOMCOZY-", na=False)]
    df = df[df["competitor_brand"] != ""]
    df = df[~df["competitor_brand"].isin(["多品牌对比", "竞品对比"])]
    df["competitor_brand"] = df["competitor_brand"].map(normalize_brand_name)
    df = df[~df["competitor_brand"].str.lower().isin(OWN_BRANDS)]
    if df.empty:
        return {}
    counter = (
        df.groupby(["country", "product_line", "competitor_brand"])
        .size()
        .reset_index(name="count")
        .sort_values(["country", "product_line", "count", "competitor_brand"], ascending=[True, True, False, True])
    )
    result: dict[tuple[str, str], str] = {}
    for (country, line), grp in counter.groupby(["country", "product_line"]):
        vals = [f"{r['competitor_brand']}({int(r['count'])})" for _, r in grp.head(5).iterrows()]
        result[(country, line)] = ", ".join(vals)
    return result


def build_cross_join(top20_df: pd.DataFrame, persona_df: pd.DataFrame, cfg_df: pd.DataFrame) -> pd.DataFrame:
    top20 = top20_df[["国家", "销售额"]].drop_duplicates().rename(columns={"国家": "country", "销售额": "top20_country_sales"})
    persona = persona_df.copy()
    persona["country"] = persona["国家"].map(clean_text)
    persona["product_line"] = persona["产品品线"].map(clean_text)
    persona = persona.drop_duplicates(subset=["country", "product_line"], keep="first")
    cfg = cfg_df.copy()
    cfg["country"] = cfg["国家"].map(clean_text)
    cfg["product_line"] = cfg["产品品线"].map(clean_text)
    cfg = cfg.drop_duplicates(subset=["country", "product_line"], keep="first")
    all_lines = sorted(persona["product_line"].dropna().unique().tolist())
    scaffold = top20.assign(_key=1).merge(pd.DataFrame({"product_line": all_lines, "_key": 1}), on="_key").drop(columns="_key")
    merged = scaffold.merge(persona, on=["country", "product_line"], how="left", suffixes=("", "_persona"))
    merged = merged.merge(cfg, on=["country", "product_line"], how="left", suffixes=("", "_cfg"))
    return merged


def apply_web_supplement(country: str, product_line: str, local_brands: list[str]) -> tuple[list[str], list[dict], list[str]]:
    rows = []
    notes = []
    result = list(local_brands)
    for sup in WEB_SUPPLEMENTS:
        if sup["country"] != country:
            continue
        if product_line not in sup["product_lines"]:
            continue
        added = [b for b in sup["local_brands"] if b not in result]
        if added:
            result.extend(added)
            notes.append(sup["note"])
        for url in sup["urls"]:
            rows.append(
                {
                    "country": country,
                    "country_code": get_code(country) or "",
                    "product_line": product_line,
                    "source_type": "web_supplement",
                    "query": sup["query"],
                    "url": url,
                    "brands_added": ", ".join(added) if added else "",
                    "note": sup["note"],
                }
            )
    return dedupe_keep_order(result), rows, dedupe_keep_order(notes)


def join_values(values: list[str]) -> str:
    vals = dedupe_keep_order([clean_text(v) for v in values if clean_text(v)])
    return ", ".join(vals)


def websites_for_brands(values: list[str]) -> str:
    pairs = []
    for brand in dedupe_keep_order(values):
        url = BRAND_SITE_MAP.get(brand)
        if url:
            pairs.append(f"{brand}: {url}")
    return "\n".join(pairs)


def auto_width(ws) -> None:
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = 0
        for cell in col[:200]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 48)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(name="Arial", bold=True)
    body_font = Font(name="Arial")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        auto_width(ws)
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    parser.add_argument("--out-file", type=Path, default=None)
    args = parser.parse_args()

    top20_path = TABLES_DIR / "dim_top20_country_insight.csv"
    persona_path = TABLES_DIR / "dim_country_product_persona.csv"
    cfg_path = TABLES_DIR / "cfg_top10_country_line.csv"
    voc_path = TABLES_DIR / "dim_voc_negative_extract.csv"

    top20_df = pd.read_csv(top20_path)
    persona_df = pd.read_csv(persona_path)
    cfg_df = pd.read_csv(cfg_path)
    voc_df = pd.read_csv(voc_path)

    merged = build_cross_join(top20_df, persona_df, cfg_df)
    top20_countries = set(merged["country"].dropna())
    voc_top = build_voc_top_brands(voc_df, top20_countries)

    master_rows: list[dict] = []
    audit_rows: list[dict] = []
    source_rows: list[dict] = [
        {
            "country": "",
            "country_code": "",
            "product_line": "",
            "source_type": "project_table",
            "query": "",
            "url": str(persona_path),
            "brands_added": "",
            "note": "核心竞争品牌/型号原始来源：dim_country_product_persona.csv",
        },
        {
            "country": "",
            "country_code": "",
            "product_line": "",
            "source_type": "project_table",
            "query": "",
            "url": str(cfg_path),
            "brands_added": "",
            "note": "销售额、优先级等补充来源：cfg_top10_country_line.csv",
        },
        {
            "country": "",
            "country_code": "",
            "product_line": "",
            "source_type": "project_table",
            "query": "",
            "url": str(voc_path),
            "brands_added": "",
            "note": "VOC 高频品牌佐证来源：dim_voc_negative_extract.csv（已排除 MOMCOZY- 内部批次与自有品牌）",
        },
    ]

    for _, row in merged.sort_values(["country", "product_line"]).iterrows():
        country = clean_text(row["country"])
        product_line = clean_text(row["product_line"])
        raw_brand = clean_text(row.get("核心竞争品牌（国际/本土）")) or clean_text(row.get("核心竞争品牌（国际/本土）_cfg"))
        raw_model = clean_text(row.get("核心竞品产品名称/型号")) or clean_text(row.get("核心竞品产品名称/型号_cfg"))
        parsed, audits = classify_tokens(country, product_line, raw_brand, raw_model)
        audit_rows.extend(audits)

        local_brands, supplement_sources, supplement_notes = apply_web_supplement(country, product_line, parsed["local_brands"])
        source_rows.extend(supplement_sources)

        intl_brands = parsed["intl_brands"]
        intl_models = parsed["intl_models"]
        local_models = parsed["local_models"]
        misc_notes = parsed["misc_notes"] + supplement_notes
        gap_flag = int(not intl_brands and not local_brands and not intl_models and not local_models)
        data_quality = []
        if raw_brand or raw_model:
            data_quality.append("project")
        if voc_top.get((country, product_line)):
            data_quality.append("voc")
        if supplement_notes:
            data_quality.append("web")
        if not data_quality:
            data_quality.append("empty")

        master_rows.append(
            {
                "国家": country,
                "产品品线": product_line,
                "本土品牌": join_values(local_brands),
                "本土网站": websites_for_brands(local_brands),
                "国际品牌": join_values(intl_brands),
                "国际网站": websites_for_brands(intl_brands),
                "本土核心商品": join_values(local_models),
                "国际核心商品": join_values(intl_models),
            }
        )

    master_df = pd.DataFrame(master_rows, columns=["国家", "产品品线", "本土品牌", "本土网站", "国际品牌", "国际网站", "本土核心商品", "国际核心商品"])
    master_df = master_df.sort_values(["国家", "产品品线"]).reset_index(drop=True)

    args.out_dir.mkdir(parents=True, exist_ok=True)
    out_path = args.out_file or (args.out_dir / f"TOP20_国家x品线_竞品品牌与核心型号_{datetime.now().strftime('%Y%m%d')}.xlsx")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        master_df.to_excel(writer, sheet_name="Master", index=False)

    style_workbook(out_path)
    print(out_path)
    print(f"master_rows={len(master_df)}")
    print(f"voc_keys={len(voc_top)}")


if __name__ == "__main__":
    main()
