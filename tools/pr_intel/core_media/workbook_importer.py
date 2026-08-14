"""Two-stage preview/approve import for the PR core-media workbook."""
from __future__ import annotations

import hashlib
import io
import json
import re
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Optional, Sequence
from urllib.parse import urlparse

import openpyxl

from tools.pr_intel.core_media.contracts import stable_id
from tools.pr_intel.core_media.storage import (
    PrMediaRepository,
    TableWriteBatch,
    WriteError,
)


P0_COUNTRY = "US"
P0_LANGUAGE = "en"
_ADMIN_ROLE = "admin"
_URL_RE = re.compile(r"https?://[^\s<>]+", re.IGNORECASE)
_US_RE = re.compile(r"(?:\bUSA\b|\bUS\b|UNITED\s+STATES)", re.IGNORECASE)
_UK_RE = re.compile(r"(?:\bUK\b|\bGB\b|UNITED\s+KINGDOM)", re.IGNORECASE)

_HEADER_ALIASES = {
    "outlet": {"媒体", "media", "outlet"},
    "region": {"类型 / 区域", "类型/区域", "type / region", "region"},
    "role": {"角色定位", "role"},
    "social": {"social media", "social"},
    "editor_name": {"editor name", "编辑姓名", "编辑"},
    "editor_title": {"editor info/ position", "editor info / position", "position"},
    "linkedin": {"linkedin"},
    "author_page": {"个人简介", "author page", "bio"},
    "cooperation": {"合作形式 (cooperation type)", "cooperation type", "合作形式"},
    "owner": {"负责人 (owner)", "owner", "负责人"},
}
_MEDIA_LEVEL_FIELDS = frozenset({"outlet", "region", "role", "social"})


class WorkbookImportError(RuntimeError):
    def __init__(self, code: str, message: str) -> None:
        super().__init__(f"{code}: {message}")
        self.code = code
        self.safe_message = message


class ImportApprovalError(RuntimeError):
    pass


@dataclass(frozen=True)
class OutletCandidate:
    outlet_id: str
    canonical_name: str
    media_type: Optional[str]
    role_tags: Optional[str]
    source_row_ref: str


@dataclass(frozen=True)
class EditionCandidate:
    edition_id: str
    outlet_id: str
    country: str
    language: str
    canonical_domain: Optional[str]
    status: str
    source_row_ref: str


@dataclass(frozen=True)
class JournalistCandidate:
    journalist_id: str
    public_name: str
    public_title: Optional[str]
    identity_status: str
    source_row_ref: str


@dataclass(frozen=True)
class AffiliationCandidate:
    affiliation_id: str
    journalist_id: str
    outlet_id: str
    edition_id: str
    role: Optional[str]
    affiliation_status: str
    source_url: Optional[str]
    source_row_ref: str


@dataclass(frozen=True)
class TouchpointCandidate:
    touchpoint_id: str
    entity_type: str
    entity_id: str
    platform: str
    public_url: str
    ownership_type: str
    collection_policy: str
    access_status: str
    source_row_ref: str


@dataclass(frozen=True)
class ImportReviewItem:
    code: str
    severity: str
    entity_type: str
    entity_id: Optional[str]
    field: str
    source_row_ref: str
    message: str


@dataclass(frozen=True)
class RawImportRecord:
    import_record_id: str
    source_row_ref: str
    raw_hash: str
    entity_type: str
    entity_id: Optional[str]
    raw_record_json: str
    status: str


@dataclass(frozen=True)
class ImportBatch:
    import_version: str
    source_file_ref: str
    source_file_sha256: str
    source_sheet: str
    created_at: str
    outlets: tuple[OutletCandidate, ...]
    editions: tuple[EditionCandidate, ...]
    journalists: tuple[JournalistCandidate, ...]
    affiliations: tuple[AffiliationCandidate, ...]
    touchpoints: tuple[TouchpointCandidate, ...]
    raw_records: tuple[RawImportRecord, ...]
    review_items: tuple[ImportReviewItem, ...]
    blank_editor_rows: int
    status: str = "preview"

    def review_items_by_code(self, code: str) -> tuple[ImportReviewItem, ...]:
        return tuple(item for item in self.review_items if item.code == code)

    def to_preview_dict(self) -> dict[str, object]:
        return {
            "import_version": self.import_version,
            "source_file_ref": self.source_file_ref,
            "source_file_sha256": self.source_file_sha256,
            "source_sheet": self.source_sheet,
            "status": self.status,
            "created_at": self.created_at,
            "counts": {
                "outlets": len(self.outlets),
                "editions": len(self.editions),
                "journalists": len(self.journalists),
                "affiliations": len(self.affiliations),
                "touchpoints": len(self.touchpoints),
                "blank_editor_rows": self.blank_editor_rows,
                "review_items": len(self.review_items),
            },
            "outlets": [asdict(item) for item in self.outlets],
            "editions": [asdict(item) for item in self.editions],
            "journalists": [asdict(item) for item in self.journalists],
            "review_items": [asdict(item) for item in self.review_items],
        }


@dataclass(frozen=True)
class ImportReport:
    import_version: str
    status: str
    outlet_count: int
    journalist_count: int
    touchpoint_count: int
    errors: tuple[WriteError, ...]


def _clean(value: object) -> Optional[str]:
    if value is None:
        return None
    text = " ".join(str(value).strip().split())
    return text or None


def _normalize_header(value: object) -> str:
    return " ".join(str(value or "").strip().casefold().split())


def _header_map(sheet: openpyxl.worksheet.worksheet.Worksheet) -> tuple[int, dict[str, int]]:
    for row_number in range(1, min(sheet.max_row, 20) + 1):
        normalized = {
            column: _normalize_header(sheet.cell(row=row_number, column=column).value)
            for column in range(1, sheet.max_column + 1)
        }
        mapped: dict[str, int] = {}
        for field, aliases in _HEADER_ALIASES.items():
            for column, header in normalized.items():
                if header in aliases:
                    mapped[field] = column
                    break
        if "outlet" in mapped and "editor_name" in mapped:
            missing = sorted(set(_HEADER_ALIASES) - set(mapped))
            if missing:
                raise WorkbookImportError(
                    "workbook_columns_missing",
                    "Required columns are missing: " + ", ".join(missing),
                )
            return row_number, mapped
    raise WorkbookImportError(
        "workbook_header_not_found",
        "Could not locate the outlet/editor header row",
    )


def _merged_media_values(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    headers: dict[str, int],
) -> dict[tuple[int, int], object]:
    media_columns = {headers[field] for field in _MEDIA_LEVEL_FIELDS}
    inherited: dict[tuple[int, int], object] = {}
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_col != merged_range.max_col:
            continue
        column = merged_range.min_col
        if column not in media_columns:
            continue
        value = sheet.cell(row=merged_range.min_row, column=column).value
        for row_number in range(merged_range.min_row, merged_range.max_row + 1):
            inherited[(row_number, column)] = value
    return inherited


def _cell_value(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    row_number: int,
    column: int,
    inherited: dict[tuple[int, int], object],
) -> object:
    cell = sheet.cell(row=row_number, column=column)
    if cell.value is not None:
        return cell.value
    return inherited.get((row_number, column))


def _extract_urls(cell: openpyxl.cell.cell.Cell) -> tuple[str, ...]:
    candidates: list[str] = []
    hyperlink = getattr(cell, "hyperlink", None)
    if hyperlink is not None and getattr(hyperlink, "target", None):
        candidates.append(str(hyperlink.target))
    for match in _URL_RE.findall(str(cell.value or "")):
        candidates.append(match.rstrip(".,;，；。)]}"))

    output: list[str] = []
    for candidate in candidates:
        parsed = urlparse(candidate)
        if parsed.scheme.casefold() in {"http", "https"} and parsed.netloc and " " not in candidate:
            if candidate not in output:
                output.append(candidate)
    return tuple(output)


def _platform(url: str, *, author_page: bool = False) -> str:
    if author_page:
        return "author_page"
    hostname = (urlparse(url).hostname or "").casefold()
    if "linkedin." in hostname:
        return "linkedin"
    if hostname == "x.com" or hostname.endswith(".x.com") or "twitter." in hostname:
        return "x"
    if "instagram." in hostname:
        return "instagram"
    if "facebook." in hostname:
        return "facebook"
    if "tiktok." in hostname:
        return "tiktok"
    if "youtube." in hostname or hostname == "youtu.be":
        return "youtube"
    return "website"


def _edition_country(region: Optional[str]) -> tuple[str, Optional[str]]:
    value = region or ""
    has_us = bool(_US_RE.search(value))
    has_uk = bool(_UK_RE.search(value))
    if has_us and has_uk:
        return P0_COUNTRY, "edition_conflict"
    if has_us:
        return P0_COUNTRY, None
    if has_uk:
        return "GB", "edition_out_of_scope"
    return "unknown", "edition_unresolved"


def _utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def preview_workbook(path: Path, sheet_name: str, import_version: str) -> ImportBatch:
    input_path = Path(path)
    if not import_version.strip():
        raise ValueError("import_version_required")
    if not input_path.is_file():
        raise FileNotFoundError(f"workbook_not_found: {input_path}")

    source_bytes = input_path.read_bytes()
    source_sha256 = hashlib.sha256(source_bytes).hexdigest()
    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(source_bytes), data_only=True, read_only=False
        )
    except Exception as exc:
        raise WorkbookImportError(
            "workbook_not_standard_xlsx",
            f"Workbook could not be parsed ({type(exc).__name__})",
        ) from exc

    try:
        if sheet_name not in workbook.sheetnames:
            raise WorkbookImportError(
                "workbook_sheet_missing", f"Sheet not found: {sheet_name}"
            )
        sheet = workbook[sheet_name]
        header_row, headers = _header_map(sheet)
        inherited = _merged_media_values(sheet, headers)

        outlets: list[OutletCandidate] = []
        editions: list[EditionCandidate] = []
        journalists: list[JournalistCandidate] = []
        affiliations: list[AffiliationCandidate] = []
        touchpoints: list[TouchpointCandidate] = []
        raw_records: list[RawImportRecord] = []
        review_items: list[ImportReviewItem] = []
        outlet_by_name: dict[str, OutletCandidate] = {}
        edition_by_outlet: dict[str, EditionCandidate] = {}
        touchpoint_keys: set[tuple[str, str, str, str]] = set()
        blank_editor_rows = 0

        for row_number in range(header_row + 1, sheet.max_row + 1):
            source_row_ref = f"{sheet_name}!{row_number}"
            values = {
                field: _clean(
                    _cell_value(sheet, row_number, column, inherited)
                    if field in _MEDIA_LEVEL_FIELDS
                    else sheet.cell(row=row_number, column=column).value
                )
                for field, column in headers.items()
            }
            raw_values = {
                field: _clean(sheet.cell(row=row_number, column=column).value)
                for field, column in headers.items()
            }
            if not any(values.values()):
                continue

            outlet_name = values["outlet"]
            if not outlet_name:
                review_items.append(
                    ImportReviewItem(
                        code="outlet_missing",
                        severity="error",
                        entity_type="row",
                        entity_id=None,
                        field="outlet",
                        source_row_ref=source_row_ref,
                        message="Row has data but is not inside an explicit merged outlet group",
                    )
                )
                continue

            outlet = outlet_by_name.get(outlet_name)
            if outlet is None:
                outlet_id = stable_id("outlet", outlet_name)
                outlet = OutletCandidate(
                    outlet_id=outlet_id,
                    canonical_name=outlet_name,
                    media_type=values["region"],
                    role_tags=values["role"],
                    source_row_ref=source_row_ref,
                )
                outlet_by_name[outlet_name] = outlet
                outlets.append(outlet)

                country, edition_issue = _edition_country(values["region"])
                edition_id = stable_id(
                    "edition", outlet_id, country, P0_LANGUAGE, "unknown"
                )
                edition = EditionCandidate(
                    edition_id=edition_id,
                    outlet_id=outlet_id,
                    country=country,
                    language=P0_LANGUAGE,
                    canonical_domain=None,
                    status="pending",
                    source_row_ref=source_row_ref,
                )
                editions.append(edition)
                edition_by_outlet[outlet_id] = edition
                if edition_issue:
                    review_items.append(
                        ImportReviewItem(
                            code=edition_issue,
                            severity="error" if edition_issue == "edition_conflict" else "warning",
                            entity_type="edition",
                            entity_id=edition_id,
                            field="region",
                            source_row_ref=source_row_ref,
                            message=(
                                "UK and US are both present; P0 keeps a pending US candidate "
                                "and does not create or merge a UK edition automatically"
                                if edition_issue == "edition_conflict"
                                else "Edition requires explicit business verification"
                            ),
                        )
                    )

                social_cell = sheet.cell(row=row_number, column=headers["social"])
                for url in _extract_urls(social_cell):
                    platform = _platform(url)
                    key = ("outlet", outlet_id, platform, url)
                    if key in touchpoint_keys:
                        continue
                    touchpoint_keys.add(key)
                    touchpoints.append(
                        TouchpointCandidate(
                            touchpoint_id=stable_id("touchpoint", *key),
                            entity_type="outlet",
                            entity_id=outlet_id,
                            platform=platform,
                            public_url=url,
                            ownership_type="official_public",
                            collection_policy="permission_pending",
                            access_status="untested",
                            source_row_ref=source_row_ref,
                        )
                    )

            editor_name = values["editor_name"]
            journalist_id: Optional[str] = None
            if not editor_name:
                blank_editor_rows += 1
                review_items.append(
                    ImportReviewItem(
                        code="editor_blank",
                        severity="info",
                        entity_type="outlet",
                        entity_id=outlet.outlet_id,
                        field="editor_name",
                        source_row_ref=source_row_ref,
                        message="Blank editor row retained as an import review item",
                    )
                )
            else:
                journalist_id = stable_id(
                    "journalist", import_version, sheet_name, str(row_number), editor_name
                )
                journalist = JournalistCandidate(
                    journalist_id=journalist_id,
                    public_name=editor_name,
                    public_title=values["editor_title"],
                    identity_status="unverified",
                    source_row_ref=source_row_ref,
                )
                journalists.append(journalist)
                if not values["editor_title"]:
                    review_items.append(
                        ImportReviewItem(
                            code="editor_title_missing",
                            severity="warning",
                            entity_type="journalist",
                            entity_id=journalist_id,
                            field="editor_title",
                            source_row_ref=source_row_ref,
                            message="Public title is missing and was not forward-filled",
                        )
                    )

                edition = edition_by_outlet[outlet.outlet_id]
                author_urls = _extract_urls(
                    sheet.cell(row=row_number, column=headers["author_page"])
                )
                affiliation = AffiliationCandidate(
                    affiliation_id=stable_id(
                        "affiliation", journalist_id, edition.edition_id, source_row_ref
                    ),
                    journalist_id=journalist_id,
                    outlet_id=outlet.outlet_id,
                    edition_id=edition.edition_id,
                    role=values["editor_title"],
                    affiliation_status="candidate",
                    source_url=author_urls[0] if author_urls else None,
                    source_row_ref=source_row_ref,
                )
                affiliations.append(affiliation)

                url_specs = (
                    ("linkedin", headers["linkedin"], False),
                    ("author_page", headers["author_page"], True),
                )
                for expected_platform, column, is_author_page in url_specs:
                    cell = sheet.cell(row=row_number, column=column)
                    for url in _extract_urls(cell):
                        platform = _platform(url, author_page=is_author_page)
                        key = ("journalist", journalist_id, platform, url)
                        if key in touchpoint_keys:
                            continue
                        touchpoint_keys.add(key)
                        touchpoints.append(
                            TouchpointCandidate(
                                touchpoint_id=stable_id("touchpoint", *key),
                                entity_type="journalist",
                                entity_id=journalist_id,
                                platform=platform,
                                public_url=url,
                                ownership_type="public_professional",
                                collection_policy=(
                                    "manual_verification_only"
                                    if expected_platform == "linkedin"
                                    else "permission_pending"
                                ),
                                access_status="untested",
                                source_row_ref=source_row_ref,
                            )
                        )

            raw_record_json = json.dumps(
                raw_values, ensure_ascii=False, sort_keys=True, separators=(",", ":")
            )
            raw_hash = hashlib.sha256(raw_record_json.encode("utf-8")).hexdigest()
            raw_records.append(
                RawImportRecord(
                    import_record_id=stable_id(
                        "import_record", import_version, sheet_name, str(row_number), raw_hash
                    ),
                    source_row_ref=source_row_ref,
                    raw_hash=raw_hash,
                    entity_type="journalist_candidate" if journalist_id else "blank_editor_row",
                    entity_id=journalist_id or outlet.outlet_id,
                    raw_record_json=raw_record_json,
                    status="preview",
                )
            )

        return ImportBatch(
            import_version=import_version,
            source_file_ref=str(input_path),
            source_file_sha256=source_sha256,
            source_sheet=sheet_name,
            created_at=_utc_now(),
            outlets=tuple(outlets),
            editions=tuple(editions),
            journalists=tuple(journalists),
            affiliations=tuple(affiliations),
            touchpoints=tuple(touchpoints),
            raw_records=tuple(raw_records),
            review_items=tuple(review_items),
            blank_editor_rows=blank_editor_rows,
        )
    finally:
        workbook.close()


def _batches_for_approval(batch: ImportBatch, approved_by_role: str) -> tuple[TableWriteBatch, ...]:
    now = _utc_now()
    source = batch.source_file_ref
    sheet = batch.source_sheet
    version = batch.import_version
    preview_ref = f"sha256:{batch.source_file_sha256}"
    return (
        TableWriteBatch(
            "ctl_import_batch",
            (
                "import_version", "source_file_ref", "source_file_sha256", "source_sheet",
                "status", "outlet_count", "journalist_count", "preview_ref",
                "approved_by_role", "approved_at", "created_at",
            ),
            ((
                version, source, batch.source_file_sha256, sheet, "approved",
                len(batch.outlets), len(batch.journalists), preview_ref,
                approved_by_role, now, batch.created_at,
            ),),
        ),
        TableWriteBatch(
            "ctl_import_record",
            (
                "import_record_id", "import_version", "source_file_ref", "source_sheet",
                "source_row_ref", "raw_hash", "entity_type", "entity_id",
                "raw_record_json", "status", "created_at",
            ),
            tuple((
                item.import_record_id, version, source, sheet, item.source_row_ref,
                item.raw_hash, item.entity_type, item.entity_id, item.raw_record_json,
                "approved", now,
            ) for item in batch.raw_records),
        ),
        TableWriteBatch(
            "dim_outlet",
            (
                "outlet_id", "canonical_name", "media_type", "role_tags_text", "status",
                "source_file_ref", "source_sheet", "source_row_ref", "import_version",
                "created_at", "updated_at",
            ),
            tuple((
                item.outlet_id, item.canonical_name, item.media_type, item.role_tags,
                "candidate", source, sheet, item.source_row_ref, version, now, now,
            ) for item in batch.outlets),
        ),
        TableWriteBatch(
            "dim_outlet_edition",
            (
                "edition_id", "outlet_id", "country", "language", "canonical_domain",
                "owner_role", "status", "verified_at", "verification_evidence_ref",
                "source_file_ref", "source_sheet", "source_row_ref", "import_version",
                "created_at", "updated_at",
            ),
            tuple((
                item.edition_id, item.outlet_id, item.country, item.language,
                item.canonical_domain, None, item.status, None, None, source, sheet,
                item.source_row_ref, version, now, now,
            ) for item in batch.editions),
        ),
        TableWriteBatch(
            "dim_journalist",
            (
                "journalist_id", "public_name", "public_title", "identity_status",
                "verified_at", "verification_evidence_ref", "source_file_ref",
                "source_sheet", "source_row_ref", "import_version", "created_at", "updated_at",
            ),
            tuple((
                item.journalist_id, item.public_name, item.public_title,
                item.identity_status, None, None, source, sheet, item.source_row_ref,
                version, now, now,
            ) for item in batch.journalists),
        ),
        TableWriteBatch(
            "bridge_journalist_affiliation",
            (
                "affiliation_id", "journalist_id", "edition_id", "role",
                "affiliation_status", "source_url", "valid_from", "valid_until",
                "verified_at", "source_file_ref", "source_sheet", "source_row_ref",
                "import_version", "created_at", "updated_at",
            ),
            tuple((
                item.affiliation_id, item.journalist_id, item.edition_id, item.role,
                item.affiliation_status, item.source_url, None, None, None, source, sheet,
                item.source_row_ref, version, now, now,
            ) for item in batch.affiliations),
        ),
        TableWriteBatch(
            "dim_touchpoint",
            (
                "touchpoint_id", "entity_type", "entity_id", "platform", "public_url",
                "ownership_type", "collection_policy", "access_status", "last_checked_at",
                "source_file_ref", "source_sheet", "source_row_ref", "import_version",
                "created_at", "updated_at",
            ),
            tuple((
                item.touchpoint_id, item.entity_type, item.entity_id, item.platform,
                item.public_url, item.ownership_type, item.collection_policy,
                item.access_status, None, source, sheet, item.source_row_ref, version,
                now, now,
            ) for item in batch.touchpoints),
        ),
    )


def approve_workbook_import(
    repository: PrMediaRepository,
    batch: ImportBatch,
    approved_by_role: str,
) -> ImportReport:
    normalized_role = approved_by_role.strip().casefold()
    if normalized_role != _ADMIN_ROLE:
        raise ImportApprovalError("admin_role_required")
    if batch.status != "preview":
        raise ImportApprovalError("preview_batch_required")

    write_report = repository.insert_table_batches(
        _batches_for_approval(batch, normalized_role)
    )
    return ImportReport(
        import_version=batch.import_version,
        status="approved" if not write_report.errors else "failed",
        outlet_count=len(batch.outlets),
        journalist_count=len(batch.journalists),
        touchpoint_count=len(batch.touchpoints),
        errors=write_report.errors,
    )
